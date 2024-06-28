import json
import os
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse
import PIL
from PIL import Image as PilImage
from pathlib import Path
PIL.Image.MAX_IMAGE_PIXELS = 108000000

def create_excel_file(json_file_path, image_file_path, excel_file_path, limit = 10, use_comment=False):
    unique_ids = set()
    with open(json_file_path, 'r') as f:
        data = json.load(f)
    # Prepare Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"

    # Define Excel headers
    headers = [
        'submissionId', 
        'submissionTitle', 
        'isTextSarcastic?', 
        'submissionImage', 
        'isImageSarcastic?', 
        'isTogetherSarcastic?'
    ]

    # Write header row
    ws.append(headers)

    if limit == None:
        limit = len(data)
    if limit < len(data):
        limit = len(data)
    for i in range(limit):
        try:
            item = data[i]
            submission_id = item.get('submissionId')
            comments = item.get("comments")
            submission_title = item.get('submissionTitle')
            submission_url = item.get('submissionURL')

            if use_comment and len(comments)>0 and comments[0].get("commentBody") != "[deleted]" and comments[0].get("commentBody") != "[removed]":
                submission_title += "\n"
                submission_title += comments[0].get("commentBody")
            

            if submission_id not in unique_ids:
                # extension = os.path.splitext(urlparse(submission_url).path)[-1]
                # if extension == ".jpg":
                #     extension = ".jpeg"
                extension = ".png"

                image_filename = f"{submission_id}{extension}"
                image_path = os.path.join(image_file_path, image_filename)
        
                is_text_sarcastic = ''
                is_image_sarcastic = ''
                is_together_sarcastic = ''

                # Add row to Excel sheet
                ws.append([
                    submission_id,
                    submission_title,
                    is_text_sarcastic,
                    "",  # Placeholder for image
                    is_image_sarcastic,
                    is_together_sarcastic
                ])

                if os.path.exists(image_path):
                    img = Image(image_path)
                    img_pil = PilImage.open(image_path)
                    aspect_ratio = img_pil.width / img_pil.height
                    img.height = 200  # Set the image height
                    img.width = int(200 * aspect_ratio)  # Set the image width maintaining aspect ratio
                    ws.add_image(img, f'D{ws.max_row}')

                    ws.row_dimensions[ws.max_row].height = 200  # Set row height
                unique_ids.add(submission_id)
        except Exception as e:
            print(f"Error while creating new excel row: {e}")


    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length+2)*1.2
        ws.column_dimensions[column].width = adjusted_width

    # Save the Excel file
    wb.save(excel_file_path)

    print(f"Excel file has been created at {excel_file_path}")

if __name__ == "__main__":
    # WARNING!! Do not forget to change these values.
    # You do not need to create excel file path manually.
    # Set LIMIT_SUBMISSION to None in order to use all the items in the JSON file.
    # e.g. LIMIT_SUBMISSION = None
    # You can only use this for categories of those subreddits whose json file is created and medias downloaded.
    SUBREDDIT_NAME = "mildlyinteresting"
    CATEGORY = "top"
    LIMIT_SUBMISSION = 200
    USE_COMMENT = True # Do u want to append comments to original text?
    #-------------------------------------------------------------------------------------
    JSON_FILE_PATH = f"json_files/{SUBREDDIT_NAME}/{CATEGORY}.json"
    IMAGE_FILE_PATH = f"media_files/{SUBREDDIT_NAME}/{CATEGORY}"
    EXCEL_FILE_DIR = f"annotated/{SUBREDDIT_NAME}"
    EXCEL_FILE_PATH = f"{EXCEL_FILE_DIR}/{CATEGORY}.xlsx"
    os.makedirs(EXCEL_FILE_DIR, exist_ok=True)

    create_excel_file(JSON_FILE_PATH, IMAGE_FILE_PATH, EXCEL_FILE_PATH, limit=LIMIT_SUBMISSION, use_comment=USE_COMMENT)